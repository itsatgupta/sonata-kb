"""Generate the Royal London Data Request workbook (.xlsx) for the consultancy team.

Source of the ask-list: 07-future-roadmap/pre-poc-readiness.md, Section 3.
Output: 07-future-roadmap/Royal-London-Data-Request.xlsx

Usage (repo root):  03-poc/agent/venv/Scripts/python.exe 07-future-roadmap/generate_data_request.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Dated filename so re-generation never collides with a file open in Excel.
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Royal-London-Data-Request-2026-08-04.xlsx")

# ---- styling helpers -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")   # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRIT_FILL = PatternFill("solid", fgColor="C00000")     # red
REQ_FILL = PatternFill("solid", fgColor="FFC000")      # amber
NICE_FILL = PatternFill("solid", fgColor="A9D08E")     # green
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREY = PatternFill("solid", fgColor="F2F2F2")

PRIORITY_STYLES = {"CRITICAL": CRIT_FILL, "Required": REQ_FILL, "Nice-to-have": NICE_FILL}

DEFAULT_LABELS = ["#", "Data requested", "Priority", "What we already have (POC)",
                  "To be filled by consultancy", "Data owner", "Target date", "Notes"]
DEFAULT_WIDTHS = [6, 50, 13, 40, 40, 20, 12, 24]


def styled_sheet(wb, title, rows, widths=None, notes=None, col_labels=None):
    """rows: (id, item, priority, already_have, ask, owner, target, note)."""
    labels = col_labels or DEFAULT_LABELS
    widths = widths or DEFAULT_WIDTHS
    ws = wb.create_sheet(title=title)
    for c, (lab, w) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=c, value=lab)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, row in enumerate(rows, start=2):
        rid, item, prio, have, ask, owner, target, note = (list(row) + [""] * 8)[:8]
        vals = [rid, item, prio, have, ask, owner, target, note]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 3 and v in PRIORITY_STYLES:
                cell.fill = PRIORITY_STYLES[v]
                cell.alignment = Alignment(vertical="top", horizontal="center")
            if c == 4:  # pre-filled 'have' column, subtle tint
                cell.fill = GREY
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(labels))}{len(rows)+1}"
    return ws


def cover_sheet(wb):
    ws = wb.active
    ws.title = "How to fill"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 110
    lines = [
        ("TITLE", "ROYAL LONDON — DATA REQUEST FOR UPGRADE-IMPACT PREPARATION"),
        ("", ""),
        ("SUB", "Prepared by: Sonata Knowledge Assistant POC · Date: 2026-08-04"),
        ("SUB", "Reference doc: 07-future-roadmap/pre-poc-readiness.md (Section 3)"),
        ("", ""),
        ("H", "PURPOSE"),
        ("P", "This workbook captures the data Royal London needs to provide so we can draft per-client "
             "upgrade-impact assessments (what changed, what is relevant to Royal London, what is risky). "
             "Fill the 'To be filled' column per row; leave rows already marked grey if we already hold the data."),
        ("", ""),
        ("H", "HOW TO FILL"),
        ("P", "1. The grey 'What we already have (POC)' column is pre-filled — do not redo it."),
        ("P", "2. 'To be filled by consultancy': add your answer; if unknown, write 'UNKNOWN'."),
        ("P", "3. 'Data owner' = the single role/person accountable for each row; 'Target date' = when it will be provided."),
        ("P", "4. Priority legend: RED = critical path (blocks the build), AMBER = required, GREEN = nice-to-have."),
        ("P", "5. Tabs A-H mirror the ask sections. Return the whole workbook when complete."),
        ("", ""),
        ("H", "SCOPE (read first)"),
        ("P", "We need 'what changed + like-for-like migration impact' only — NOT feature-adoption or product "
             "roadmap material. Please do not gather out-of-scope content."),
        ("P", "All access to Jira/Wiki/Bitbucket/X-ray is READ-ONLY (rule 2). No write-back is performed."),
        ("", ""),
        ("H", "ACCESS WE ALREADY USE IN THE POC (read-only)"),
        ("P", "Jira: defect tracker filter 90250 · CART tests filter 103721 · project RLSI"),
        ("P", "Wiki: release-notes pages 16.4 (1001572222) and 16.5 (1007867808)"),
        ("P", "Contacts so far: Pratigya (Feature 1 SME), Sanjay Joshi (Feature 2 SME)"),
        ("", ""),
        ("H", "WHAT TO RETURN"),
        ("P", "The filled workbook + a named contact for Q&A: delivery lead, account manager, technical SME, QA lead."),
        ("", ""),
        ("BY", "Please return to: [KB / Upgrade CoE lead] — [email]"),
    ]
    for r, (kind, text) in enumerate(lines, start=1):
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "TITLE":
            cell.font = Font(bold=True, size=14, color="1F4E78")
        elif kind == "H":
            cell.font = Font(bold=True, size=12, color="1F4E78")
        elif kind == "BY":
            cell.font = Font(bold=True, italic=True)
        elif kind == "SUB":
            cell.font = Font(italic=True, color="595959")
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def build():
    wb = Workbook()
    cover_sheet(wb)
    W = None  # widths default

    styled_sheet(wb, "A - Identity & Env", [
        ("A1", "Client legal entity + region + operating countries", "CRITICAL", "Royal London (EMEA/UK)", "", "Delivery lead", "", ""),
        ("A2", "Instance(s): environment URLs, number of instances, dev/UAT/prod", "Required", "", "", "Delivery lead", "", ""),
        ("A3", "Current Sonata version (production) + go-live date", "CRITICAL", "v16.1 (from phase-1-data-requirements.md; confirm)", "", "Delivery lead", "", ""),
        ("A4", "Full version history installed to date", "Required", "", "", "Delivery lead", "", ""),
        ("A5", "Support/contract tier (e.g. Premium, SLA)", "Nice-to-have", "", "", "Account manager", "", ""),
    ])

    styled_sheet(wb, "B - Versions & mapping", [
        ("B1", "Ordered list of trunk releases between two versions (e.g. 16.2 -> 16.5)", "CRITICAL", "Release-note page IDs for 16.4/16.5 available", "", "Upgrade team", "", ""),
        ("B2", "Validated fixVersion -> trunk release mapping for the client's releases", "CRITICAL", "OPEN QUESTION in 04-data-sources/jira.md", "", "Upgrade team", "", "Version-diff core (jira_version_range)"),
        ("B3", "Planned/committed upgrade version + target quarter", "Required", "UAR feature scheduled c/w v16.2 (Q4)", "", "Upgrade team", "", ""),
    ])

    styled_sheet(wb, "C - Modules & usage", [
        ("C1", "Licensed modules/components list (or entitlement/contract export)", "CRITICAL", "UAR, ISA, Payroll, GL", "", "Contract manager", "", ""),
        ("C2", "Which modules are ACTUALLY in production use vs licensed", "CRITICAL", "", "", "Delivery lead", "", "Often differs from licensed; drives relevance"),
        ("C3", "Does a client config/entitlement repository already exist to ingest? (vs build fresh)", "Required", "", "", "Delivery lead", "", "gap-analysis question 2"),
        ("C4", "Config/parameter exports (anonymised if required)", "Required", "", "", "Delivery lead", "", "Safe-to-share / redact secrets"),
    ])

    styled_sheet(wb, "D - Customisations", [
        ("D1", "Change-request / custom-code inventory (Jira project keys)", "CRITICAL", "Project RLSI", "", "Upgrade team", "", "Highest risk area"),
        ("D2", "Which Sonata components carry client-specific code/config (Bitbucket repos/modules)", "CRITICAL", "Needs path->module map (WS2)", "", "Engineering", "", ""),
        ("D3", "Prior-upgrade documents: what was customised last time", "Required", "", "", "Upgrade team", "", ""),
        ("D4", "Known deviations from stock behaviour", "Required", "", "", "Delivery lead", "", ""),
    ])

    styled_sheet(wb, "E - Regulatory", [
        ("E1", "Regulatory jurisdictions affecting this client (UK/EU for EMEA)", "Required", "Royal London = UK (PRIIPs/RDR/ISA)", "", "Compliance/Product", "", ""),
        ("E2", "Region-specific rules the assistant must not generalise from", "Required", "", "", "Compliance/Product", "", "Flag as data-gaps, never guess (rule 4)"),
    ])

    styled_sheet(wb, "F - Defects & history", [
        ("F1", "Prior version-to-version jumps + issues encountered", "Required", "Defect tracker project RLSI; Jira filter 90250", "", "Upgrade team", "", ""),
        ("F2", "Historical defect density per module/version", "Nice-to-have", "", "", "QA lead", "", "Feeds risk scoring"),
        ("F3", "Known open blockers on the current version", "Required", "", "", "Delivery lead", "", ""),
    ])

    styled_sheet(wb, "G - Test coverage", [
        ("G1", "CART test sets per module/version", "Required", "Jira filter 103721", "", "QA lead", "", ""),
        ("G2", "X-ray test execution status per module", "Nice-to-have", "Phase 2 scope; xray_search stubbed", "", "QA lead", "", ""),
    ])

    styled_sheet(wb, "H - Contacts & access", [
        ("H1", "Named roles: delivery lead, account manager, technical SME, QA lead", "CRITICAL", "SMEs: Pratigya, Sanjay Joshi", "", "Consultancy PM", "", ""),
        ("H2", "Access granted (read-only) to Jira/Wiki/Bitbucket/X-ray for this client", "Required", "Jira 90250/103721/RLSI; Wiki 1001572222/1007867808", "", "Consultancy PM/Security", "", "Confirm list is complete for Royal London"),
        ("H3", "Authorisation path / do-not-contact list for client-specific data", "Required", "", "", "Consultancy PM", "", ""),
    ])

    wb.save(OUT)
    print("Wrote:", OUT)


if __name__ == "__main__":
    build()